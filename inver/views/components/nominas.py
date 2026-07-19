import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime, timedelta
from ...models import RegistroProduccion


class NominasView:
    def _mostrar_nomina_semanal(self):
        self._limpiar_contenido()
        frame = tk.Frame(self.content_area, bg=self.colors['bg_main'])
        frame.pack(fill='both', expand=True)
        self.current_frame = frame

        header = tk.Frame(frame, bg=self.colors['bg_main'], height=50)
        header.pack(fill='x', padx=15, pady=(8,4))
        header.pack_propagate(False)
        header.config(highlightbackground=self.colors['border'], highlightthickness=1)
        header_inner = tk.Frame(header, bg=self.colors['bg_main'])
        header_inner.place(relx=0.5, rely=0.5, anchor='center')
        tk.Label(header_inner, text="💰 Nómina", font=('Segoe UI', 16, 'bold'),
                 bg=self.colors['bg_main'], fg=self.colors['accent']).pack()

        main_panel = tk.Frame(frame, bg=self.colors['bg_main'])
        main_panel.pack(fill='both', expand=True, padx=12, pady=4)

        sel_frame = tk.Frame(main_panel, bg=self.colors['card_bg'])
        sel_frame.pack(fill='x', pady=(0,4))
        sel_frame.config(highlightbackground=self.colors['border'], highlightthickness=1)

        tk.Label(sel_frame, text="Trabajador:", font=('Segoe UI', 8, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['text_light']).pack(side='left', padx=(6,3), pady=4)
        self.combo_nomina_trabajador = ttk.Combobox(sel_frame, font=('Segoe UI', 9), width=30)
        self.combo_nomina_trabajador.pack(side='left', padx=(0,8), pady=4)
        self.combo_nomina_trabajador.bind('<<ComboboxSelected>>', self._cargar_produccion_trabajador)

        tk.Label(sel_frame, text="Cargo:", font=('Segoe UI', 8, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['text_light']).pack(side='left', padx=(6,3), pady=4)
        self.label_cargo_nomina = tk.Label(sel_frame, text="", font=('Segoe UI', 9, 'bold'),
                                           bg=self.colors['card_bg'], fg=self.colors['accent'])
        self.label_cargo_nomina.pack(side='left', pady=4)

        form_frame = tk.LabelFrame(main_panel, text="Registro de Producción", font=('Segoe UI', 9, 'bold'),
                                   bg=self.colors['card_bg'], fg=self.colors['accent'], bd=2, relief='groove')
        form_frame.pack(fill='x', pady=2)

        row1 = tk.Frame(form_frame, bg=self.colors['card_bg'])
        row1.pack(fill='x', padx=6, pady=2)
        for i in range(4):
            row1.grid_columnconfigure(i, weight=1)

        f1 = tk.Frame(row1, bg=self.colors['card_bg'])
        f1.grid(row=0, column=0, sticky='ew', padx=2)
        tk.Label(f1, text="Fecha:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        fecha_frame = tk.Frame(f1, bg=self.colors['card_bg'])
        fecha_frame.pack(fill='x')
        self.entry_fecha_prod = tk.Entry(fecha_frame, font=('Segoe UI', 9),
                                         bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                         insertbackground=self.colors['text_light'],
                                         state='readonly', readonlybackground=self.colors['input_bg'],
                                         highlightbackground=self.colors['input_border'],
                                         highlightthickness=1, relief='solid')
        self.entry_fecha_prod.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_fecha_prod.pack(side='left', fill='x', expand=True)
        tk.Button(fecha_frame, text='📅', width=6, height=1, command=lambda: self._abrir_calendario(self.entry_fecha_prod),
                  bg=self.colors['accent'], fg='white', relief='flat', font=('Segoe UI', 7),
                  anchor='center', justify='center').pack(side='left', padx=(1,0))

        f2 = tk.Frame(row1, bg=self.colors['card_bg'])
        f2.grid(row=0, column=1, sticky='ew', padx=2)
        tk.Label(f2, text="Ticket:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_ticket_prod = tk.Entry(f2, font=('Segoe UI', 9),
                                          bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                          insertbackground=self.colors['text_light'],
                                          highlightbackground=self.colors['input_border'],
                                          highlightthickness=1, relief='solid')
        self.entry_ticket_prod.pack(fill='x')

        f3 = tk.Frame(row1, bg=self.colors['card_bg'])
        f3.grid(row=0, column=2, sticky='ew', padx=2)
        tk.Label(f3, text="Referencia:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_referencia_prod = tk.Entry(f3, font=('Segoe UI', 9),
                                              bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                              insertbackground=self.colors['text_light'],
                                              highlightbackground=self.colors['input_border'],
                                              highlightthickness=1, relief='solid')
        self.entry_referencia_prod.pack(fill='x')

        f4 = tk.Frame(row1, bg=self.colors['card_bg'])
        f4.grid(row=0, column=3, sticky='ew', padx=2)
        tk.Label(f4, text="Color:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_color_prod = tk.Entry(f4, font=('Segoe UI', 9),
                                         bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                         insertbackground=self.colors['text_light'],
                                         highlightbackground=self.colors['input_border'],
                                         highlightthickness=1, relief='solid',
                                         validate='key', validatecommand=(self.vcmd_letras, '%P'))
        self.entry_color_prod.pack(fill='x')

        row2 = tk.Frame(form_frame, bg=self.colors['card_bg'])
        row2.pack(fill='x', padx=6, pady=2)
        for i in range(4):
            row2.grid_columnconfigure(i, weight=1)

        f5 = tk.Frame(row2, bg=self.colors['card_bg'])
        f5.grid(row=0, column=0, sticky='ew', padx=2)
        tk.Label(f5, text="Pares:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_pares_prod = tk.Entry(f5, font=('Segoe UI', 9),
                                         bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                         insertbackground=self.colors['text_light'],
                                         highlightbackground=self.colors['input_border'],
                                         highlightthickness=1, relief='solid',
                                         validate='key', validatecommand=(self.vcmd_numeros, '%P'))
        self.entry_pares_prod.pack(fill='x')
        self.entry_pares_prod.bind('<KeyRelease>', self._actualizar_total_prod)

        f6 = tk.Frame(row2, bg=self.colors['card_bg'])
        f6.grid(row=0, column=1, sticky='ew', padx=2)
        tk.Label(f6, text="Pedido:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_pedido_prod = tk.Entry(f6, font=('Segoe UI', 9),
                                          bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                          insertbackground=self.colors['text_light'],
                                          highlightbackground=self.colors['input_border'],
                                          highlightthickness=1, relief='solid',
                                          validate='key', validatecommand=(self.vcmd_numeros, '%P'))
        self.entry_pedido_prod.pack(fill='x')
        self.entry_pedido_prod.bind('<KeyRelease>', self._actualizar_total_prod)

        f7 = tk.Frame(row2, bg=self.colors['card_bg'])
        f7.grid(row=0, column=2, sticky='ew', padx=2)
        tk.Label(f7, text="Precio:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_precio_prod = tk.Entry(f7, font=('Segoe UI', 9),
                                          bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                          insertbackground=self.colors['text_light'],
                                          highlightbackground=self.colors['input_border'],
                                          highlightthickness=1, relief='solid',
                                          validate='key', validatecommand=(self.vcmd_decimal, '%P'))
        self.entry_precio_prod.pack(fill='x')
        self.entry_precio_prod.bind('<KeyRelease>', self._actualizar_total_prod)

        f8 = tk.Frame(row2, bg=self.colors['card_bg'])
        f8.grid(row=0, column=3, sticky='ew', padx=2)
        tk.Label(f8, text="Total:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).pack(anchor='w')
        self.entry_total_prod = tk.Entry(f8, font=('Segoe UI', 9),
                                         bg=self.colors['input_bg'], fg=self.colors['text_light'],
                                         insertbackground=self.colors['text_light'],
                                         state='readonly', readonlybackground=self.colors['input_bg'],
                                         highlightbackground=self.colors['input_border'],
                                         highlightthickness=1, relief='solid')
        self.entry_total_prod.pack(fill='x')

        ttk.Separator(main_panel, orient='horizontal').pack(fill='x', pady=4)

        table_frame = tk.Frame(main_panel, bg=self.colors['bg_main'])
        table_frame.pack(fill='both', expand=True, pady=2)

        scroll_y = ttk.Scrollbar(table_frame)
        scroll_y.pack(side='right', fill='y')
        scroll_x = ttk.Scrollbar(table_frame, orient='horizontal')
        scroll_x.pack(side='bottom', fill='x')

        self.tree_produccion = ttk.Treeview(table_frame,
                                            columns=('Fecha', 'Ticket', 'Referencia', 'Color', 'Pares', 'Pedido', 'Precio', 'Total'),
                                            show='headings', height=5,
                                            yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        scroll_y.config(command=self.tree_produccion.yview)
        scroll_x.config(command=self.tree_produccion.xview)

        cols = [('Fecha', 90), ('Ticket', 70), ('Referencia', 100), ('Color', 100),
                ('Pares', 50), ('Pedido', 50), ('Precio', 100), ('Total', 100)]
        for col, w in cols:
            self.tree_produccion.heading(col, text=col)
            self.tree_produccion.column(col, width=w, anchor='center')
        self.tree_produccion.pack(fill='both', expand=True)
        self.tree_produccion.bind('<<TreeviewSelect>>', self._cargar_datos_produccion_seleccion)

        total_frame = tk.Frame(main_panel, bg=self.colors['card_bg'])
        total_frame.pack(fill='x', pady=2)
        total_frame.config(highlightbackground=self.colors['border'], highlightthickness=1)
        self.label_total_general = tk.Label(total_frame, text="Total General: $0,00",
                                            font=('Segoe UI', 9, 'bold'),
                                            bg=self.colors['card_bg'], fg=self.colors['accent'])
        self.label_total_general.pack(pady=2, padx=6, anchor='e')

        btn_frame = tk.Frame(main_panel, bg=self.colors['bg_main'])
        btn_frame.pack(fill='x', pady=4)

        self._crear_boton(btn_frame, "AGREGAR", self._agregar_produccion, 'success').pack(side='left', padx=2)
        self._crear_boton(btn_frame, "MODIFICAR", self._modificar_produccion, 'warning').pack(side='left', padx=2)
        self._crear_boton(btn_frame, "ELIMINAR", self._eliminar_produccion, 'danger').pack(side='left', padx=2)
        self._crear_boton(btn_frame, "CALCULAR NÓMINA", self._abrir_calculo_nomina, 'info').pack(side='left', padx=2)
        self._crear_boton(btn_frame, "📊 EXPORTAR EXCEL", self._exportar_produccion_excel, 'info').pack(side='left', padx=2)

        self._actualizar_listas_combos()

    def _exportar_produccion_excel(self):
        """Exporta los registros de producción visibles en la tabla a un archivo Excel."""
        seleccion = self.combo_nomina_trabajador.get()
        if not seleccion:
            self._mostrar_mensaje("⚠️ Seleccione un trabajador para exportar", 'warning')
            return

        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            self._mostrar_mensaje("La librería openpyxl no está instalada. Instálala con: pip install openpyxl", 'error')
            return

        items = self.tree_produccion.get_children()
        if not items:
            self._mostrar_mensaje("No hay registros de producción para exportar", 'warning')
            return

        cedula = seleccion.split(" - ")[0]
        trabajador = self.nomina.trabajadores.get(cedula)
        if not trabajador:
            self._mostrar_mensaje("Trabajador no encontrado", 'error')
            return

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar producción de trabajador",
            initialfile=f"Produccion_{trabajador.nombres}_{trabajador.apellidos}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        )
        if not archivo:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Producción"

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            center_alignment = Alignment(horizontal='center', vertical='center')
            money_alignment = Alignment(horizontal='right', vertical='center')
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Título
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = f"INVERSPORT - REGISTRO DE PRODUCCIÓN"
            title_cell.font = Font(size=14, bold=True)
            title_cell.alignment = Alignment(horizontal='center')
            
            ws.merge_cells('A2:H2')
            info_cell = ws['A2']
            info_cell.value = f"Trabajador: {trabajador.nombre} | Cédula: {cedula} | Cargo: {trabajador.cargo}"
            info_cell.font = Font(size=11)
            info_cell.alignment = Alignment(horizontal='center')

            ws.merge_cells('A3:H3')
            fecha_cell = ws['A3']
            fecha_cell.value = f"Fecha de exportación: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            fecha_cell.font = Font(size=10)
            fecha_cell.alignment = Alignment(horizontal='center')

            # Encabezados
            headers = ['Fecha', 'Ticket', 'Referencia', 'Color', 'Pares', 'Pedido', 'Precio', 'Total']
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            # Datos
            total_general = 0
            for row_idx, item in enumerate(items, start=6):
                values = self.tree_produccion.item(item, 'values')
                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    
                    # Formato de moneda para Precio y Total
                    if col_idx == 7 or col_idx == 8:
                        try:
                            valor_limpio = value.replace('$', '').replace('.', '').replace(',', '.')
                            cell.value = float(valor_limpio)
                            cell.number_format = '#,##0.00 "Bs"'
                            cell.alignment = money_alignment
                            if col_idx == 8:
                                total_general += float(valor_limpio)
                        except:
                            pass

            # Total general
            row_total = 6 + len(items)
            ws.merge_cells(f'A{row_total}:G{row_total}')
            total_label = ws.cell(row=row_total, column=1, value="TOTAL GENERAL")
            total_label.font = Font(bold=True)
            total_label.alignment = Alignment(horizontal='right')
            total_label.border = thin_border
            
            total_cell = ws.cell(row=row_total, column=8, value=total_general)
            total_cell.font = Font(bold=True)
            total_cell.number_format = '#,##0.00 "Bs"'
            total_cell.alignment = money_alignment
            total_cell.border = thin_border

            # Ajustar anchos de columna
            for col_idx in range(1, 9):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 15
            ws.column_dimensions['B'].width = 12  # Ticket más angosto
            ws.column_dimensions['D'].width = 12  # Color
            ws.column_dimensions['E'].width = 10  # Pares
            ws.column_dimensions['F'].width = 10  # Pedido

            wb.save(archivo)
            self._mostrar_mensaje(f"Producción exportada correctamente a {archivo}", 'success')
            
        except Exception as e:
            self._mostrar_mensaje(f"Error al exportar: {str(e)}", 'error')

    def _cargar_produccion_trabajador(self, event=None):
        seleccion = self.combo_nomina_trabajador.get()
        if not seleccion:
            return
        cedula = seleccion.split(" - ")[0]
        trabajador = self.nomina.trabajadores.get(cedula)
        if trabajador:
            self.label_cargo_nomina.config(text=trabajador.cargo)
        else:
            self.label_cargo_nomina.config(text="")
        for item in self.tree_produccion.get_children():
            self.tree_produccion.delete(item)
        registros = self.nomina.obtener_produccion(cedula)
        for reg in registros:
            self.tree_produccion.insert('', 'end', values=(
                reg.fecha, reg.ticket, reg.referencia, reg.color,
                reg.pares, reg.pedido,
                self._formato_cop(reg.precio),
                self._formato_cop(reg.total)
            ))
        self._limpiar_formulario_produccion()
        self._actualizar_total_general()

    def _actualizar_total_prod(self, event=None):
        try:
            pares = float(self.entry_pares_prod.get()) if self.entry_pares_prod.get() else 0
            precio = float(self.entry_precio_prod.get()) if self.entry_precio_prod.get() else 0
            total = pares * precio
            self.entry_total_prod.config(state='normal')
            self.entry_total_prod.delete(0, tk.END)
            self.entry_total_prod.insert(0, self._formato_cop(total))
            self.entry_total_prod.config(state='readonly')
        except ValueError:
            pass

    def _actualizar_total_general(self):
        total = 0.0
        for item in self.tree_produccion.get_children():
            values = self.tree_produccion.item(item, 'values')
            if values and len(values) >= 8:
                try:
                    total += self._parsear_cop(values[7])
                except:
                    pass
        self.label_total_general.config(text=f"Total General: {self._formato_cop(total)}")

    def _agregar_produccion(self):
        seleccion = self.combo_nomina_trabajador.get()
        if not seleccion:
            self._mostrar_mensaje("Seleccione un trabajador", 'warning')
            return
        cedula = seleccion.split(" - ")[0]
        fecha = self.entry_fecha_prod.get()
        ticket = self.entry_ticket_prod.get().strip()
        referencia = self.entry_referencia_prod.get().strip()
        color = self.entry_color_prod.get().strip()
        pares_str = self.entry_pares_prod.get().strip()
        pedido_str = self.entry_pedido_prod.get().strip()
        precio_str = self.entry_precio_prod.get().strip()
        if not all([fecha, ticket, referencia, color, pares_str, pedido_str, precio_str]):
            self._mostrar_mensaje("Complete todos los campos", 'warning')
            return
        try:
            pares = int(pares_str)
            if pares <= 0:
                self._mostrar_mensaje("Los pares deben ser un número positivo", 'warning')
                return
        except ValueError:
            self._mostrar_mensaje("Pares debe ser un número entero", 'error')
            return
        try:
            pedido = int(pedido_str)
            if pedido <= 0:
                self._mostrar_mensaje("El pedido debe ser un número positivo", 'warning')
                return
        except ValueError:
            self._mostrar_mensaje("Pedido debe ser un número entero", 'error')
            return
        try:
            precio = float(precio_str)
            if precio <= 0:
                self._mostrar_mensaje("El precio debe ser un número positivo", 'warning')
                return
        except ValueError:
            self._mostrar_mensaje("Precio debe ser un número (puede usar punto decimal)", 'error')
            return
        nuevo = RegistroProduccion(fecha, ticket, referencia, color, pares, pedido, precio)
        self.nomina.agregar_produccion(cedula, nuevo)
        self._cargar_produccion_trabajador()
        self._mostrar_mensaje(f"Registro agregado (Ticket: {ticket})", 'success')

    def _modificar_produccion(self):
        seleccion = self.tree_produccion.selection()
        if not seleccion:
            self._mostrar_mensaje("Seleccione un registro para modificar", 'warning')
            return
        item = seleccion[0]
        indice = self.tree_produccion.index(item)
        trabajador_sel = self.combo_nomina_trabajador.get()
        if not trabajador_sel:
            return
        cedula = trabajador_sel.split(" - ")[0]
        fecha = self.entry_fecha_prod.get()
        ticket = self.entry_ticket_prod.get().strip()
        referencia = self.entry_referencia_prod.get().strip()
        color = self.entry_color_prod.get().strip()
        pares_str = self.entry_pares_prod.get().strip()
        pedido_str = self.entry_pedido_prod.get().strip()
        precio_str = self.entry_precio_prod.get().strip()
        if not all([fecha, ticket, referencia, color, pares_str, pedido_str, precio_str]):
            self._mostrar_mensaje("Complete todos los campos", 'warning')
            return
        try:
            pares = int(pares_str)
            if pares <= 0:
                self._mostrar_mensaje("Los pares deben ser un número positivo", 'warning')
                return
        except ValueError:
            self._mostrar_mensaje("Pares debe ser un número entero", 'error')
            return
        try:
            pedido = int(pedido_str)
            if pedido <= 0:
                self._mostrar_mensaje("El pedido debe ser un número positivo", 'warning')
                return
        except ValueError:
            self._mostrar_mensaje("Pedido debe ser un número entero", 'error')
            return
        try:
            precio = float(precio_str)
            if precio <= 0:
                self._mostrar_mensaje("El precio debe ser un número positivo", 'warning')
                return
        except ValueError:
            self._mostrar_mensaje("Precio debe ser un número (puede usar punto decimal)", 'error')
            return
        nuevo = RegistroProduccion(fecha, ticket, referencia, color, pares, pedido, precio)
        if self.nomina.modificar_produccion(cedula, indice, nuevo):
            self._cargar_produccion_trabajador()
            self._mostrar_mensaje(f"Registro modificado (Ticket: {ticket})", 'success')
        else:
            self._mostrar_mensaje("Error al modificar", 'error')

    def _eliminar_produccion(self):
        seleccion = self.tree_produccion.selection()
        if not seleccion:
            self._mostrar_mensaje("Seleccione un registro para eliminar", 'warning')
            return
        if not messagebox.askyesno("Confirmar", "¿Eliminar este registro de producción?"):
            return
        item = seleccion[0]
        indice = self.tree_produccion.index(item)
        trabajador_sel = self.combo_nomina_trabajador.get()
        if not trabajador_sel:
            return
        cedula = trabajador_sel.split(" - ")[0]
        if self.nomina.eliminar_produccion(cedula, indice):
            self._cargar_produccion_trabajador()
            self._mostrar_mensaje("Registro eliminado", 'success')
        else:
            self._mostrar_mensaje("Error al eliminar", 'error')

    def _limpiar_formulario_produccion(self):
        self.entry_color_prod.delete(0, tk.END)
        self.entry_ticket_prod.delete(0, tk.END)
        self.entry_referencia_prod.delete(0, tk.END)
        self.entry_pares_prod.delete(0, tk.END)
        self.entry_pedido_prod.delete(0, tk.END)
        self.entry_precio_prod.delete(0, tk.END)
        self.entry_total_prod.config(state='normal')
        self.entry_total_prod.delete(0, tk.END)
        self.entry_total_prod.config(state='readonly')
        self.entry_fecha_prod.config(state='normal')
        self.entry_fecha_prod.delete(0, tk.END)
        self.entry_fecha_prod.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.entry_fecha_prod.config(state='readonly')
        for item in self.tree_produccion.selection():
            self.tree_produccion.selection_remove(item)

    def _cargar_datos_produccion_seleccion(self, event=None):
        seleccion = self.tree_produccion.selection()
        if not seleccion:
            return
        item = seleccion[0]
        valores = self.tree_produccion.item(item, 'values')
        if not valores:
            return
        self.entry_fecha_prod.config(state='normal')
        self.entry_fecha_prod.delete(0, tk.END)
        self.entry_fecha_prod.insert(0, valores[0])
        self.entry_fecha_prod.config(state='readonly')
        self.entry_ticket_prod.delete(0, tk.END)
        self.entry_ticket_prod.insert(0, valores[1])
        self.entry_referencia_prod.delete(0, tk.END)
        self.entry_referencia_prod.insert(0, valores[2])
        self.entry_color_prod.delete(0, tk.END)
        self.entry_color_prod.insert(0, valores[3])
        self.entry_pares_prod.delete(0, tk.END)
        self.entry_pares_prod.insert(0, valores[4])
        self.entry_pedido_prod.delete(0, tk.END)
        self.entry_pedido_prod.insert(0, valores[5])
        precio_parseado = self._parsear_cop(valores[6])
        self.entry_precio_prod.delete(0, tk.END)
        self.entry_precio_prod.insert(0, f"{precio_parseado:.2f}")
        self._actualizar_total_prod()

    def _abrir_calculo_nomina(self):
        top = tk.Toplevel(self.root)
        top.title("Cálculo de Nómina")
        top.geometry("700x750")
        top.configure(bg=self.colors['bg_dark'])
        top.grab_set()
        top.transient(self.root)

        periodo_var = tk.StringVar(value="Quincenal")
        salario_minimo_var = tk.StringVar(value="130.00")
        cestaticket_diario_var = tk.StringVar(value="15.00")
        lopcymat_porc_var = tk.StringVar(value="1.0")
        ivss_patronal_porc_var = tk.StringVar(value="11.0")
        ivss_trab_porc_var = tk.StringVar(value="4.0")
        faov_porc_var = tk.StringVar(value="1.0")
        inces_porc_var = tk.StringVar(value="0.5")

        main_frame = tk.Frame(top, bg=self.colors['bg_main'])
        main_frame.pack(fill='both', expand=True, padx=15, pady=15)

        tk.Label(main_frame, text="💰 Cálculo de Nómina", font=('Segoe UI', 16, 'bold'),
                 bg=self.colors['bg_main'], fg=self.colors['accent']).pack(pady=(0,10))

        sel_frame = tk.Frame(main_frame, bg=self.colors['card_bg'])
        sel_frame.pack(fill='x', pady=5)
        sel_frame.config(highlightbackground=self.colors['border'], highlightthickness=1)

        tk.Label(sel_frame, text="Trabajador:", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['text_light']).pack(side='left', padx=(6,3), pady=4)
        combo_trab = ttk.Combobox(sel_frame, font=('Segoe UI', 9), width=30)
        combo_trab.pack(side='left', padx=(0,8), pady=4)
        combo_trab['values'] = [f"{cedula} - {t.nombre}" for cedula, t in self.nomina.trabajadores.items()]

        params_frame = tk.LabelFrame(main_frame, text="Parámetros de Cálculo", font=('Segoe UI', 10, 'bold'),
                                     bg=self.colors['card_bg'], fg=self.colors['accent'], bd=2, relief='groove')
        params_frame.pack(fill='x', pady=8)

        params_inner = tk.Frame(params_frame, bg=self.colors['card_bg'])
        params_inner.pack(padx=10, pady=8, fill='x')

        for i in range(4):
            params_inner.columnconfigure(i, weight=1)

        tk.Label(params_inner, text="Período:", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=0, column=0, sticky='w', padx=2)
        combo_periodo = ttk.Combobox(params_inner, textvariable=periodo_var, values=['Quincenal', 'Mensual'], width=12, font=('Segoe UI', 8))
        combo_periodo.grid(row=1, column=0, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="Salario Mínimo (Bs):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=0, column=1, sticky='w', padx=2)
        entry_sal_min = tk.Entry(params_inner, textvariable=salario_minimo_var, width=12, font=('Segoe UI', 8),
                                 bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_sal_min.grid(row=1, column=1, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="Cestaticket diario (Bs):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=0, column=2, sticky='w', padx=2)
        entry_cestaticket = tk.Entry(params_inner, textvariable=cestaticket_diario_var, width=12, font=('Segoe UI', 8),
                                     bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_cestaticket.grid(row=1, column=2, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="LOPCYMAT (%):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=0, column=3, sticky='w', padx=2)
        entry_lopcymat = tk.Entry(params_inner, textvariable=lopcymat_porc_var, width=12, font=('Segoe UI', 8),
                                  bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_lopcymat.grid(row=1, column=3, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="IVSS Patronal (%):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=2, column=0, sticky='w', padx=2)
        entry_ivss_pat = tk.Entry(params_inner, textvariable=ivss_patronal_porc_var, width=12, font=('Segoe UI', 8),
                                  bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_ivss_pat.grid(row=3, column=0, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="IVSS Trabajador (%):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=2, column=1, sticky='w', padx=2)
        entry_ivss_trab = tk.Entry(params_inner, textvariable=ivss_trab_porc_var, width=12, font=('Segoe UI', 8),
                                   bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_ivss_trab.grid(row=3, column=1, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="FAOV (%):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=2, column=2, sticky='w', padx=2)
        entry_faov = tk.Entry(params_inner, textvariable=faov_porc_var, width=12, font=('Segoe UI', 8),
                              bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_faov.grid(row=3, column=2, sticky='ew', padx=2, pady=2)

        tk.Label(params_inner, text="INCES (%):", font=('Segoe UI', 8), bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=2, column=3, sticky='w', padx=2)
        entry_inces = tk.Entry(params_inner, textvariable=inces_porc_var, width=12, font=('Segoe UI', 8),
                               bg=self.colors['input_bg'], fg=self.colors['text_light'])
        entry_inces.grid(row=3, column=3, sticky='ew', padx=2, pady=2)

        btn_calcular = self._crear_boton(main_frame, "CALCULAR", 
                                         lambda: self._procesar_calculo_nomina(
                                             top, combo_trab, periodo_var, salario_minimo_var,
                                             cestaticket_diario_var, lopcymat_porc_var,
                                             ivss_patronal_porc_var, ivss_trab_porc_var,
                                             faov_porc_var, inces_porc_var
                                         ), 'primary')
        btn_calcular.pack(pady=8)

        resultados_frame = tk.LabelFrame(main_frame, text="Resultados", font=('Segoe UI', 10, 'bold'),
                                         bg=self.colors['card_bg'], fg=self.colors['accent'], bd=2, relief='groove')
        resultados_frame.pack(fill='both', expand=True, pady=5)

        results_inner = tk.Frame(resultados_frame, bg=self.colors['card_bg'])
        results_inner.pack(padx=10, pady=8, fill='both', expand=True)

        results_inner.columnconfigure(0, weight=1)
        results_inner.columnconfigure(1, weight=1)

        tk.Label(results_inner, text="ASIGNACIONES", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['accent']).grid(row=0, column=0, columnspan=2, sticky='w', pady=(0,4))

        asignaciones = [
            ("Salario de producción", "0,00"),
            ("Cestaticket", "0,00"),
            ("Alícuota utilidades", "0,00"),
            ("Alícuota bono vacacional", "0,00"),
        ]
        labels_asig = {}
        for i, (text, _) in enumerate(asignaciones):
            tk.Label(results_inner, text=f"{text}:", font=('Segoe UI', 8),
                     bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=i+1, column=0, sticky='w', pady=1)
            lbl = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 8, 'bold'),
                           bg=self.colors['card_bg'], fg=self.colors['success'])
            lbl.grid(row=i+1, column=1, sticky='e', pady=1)
            labels_asig[text] = lbl

        tk.Label(results_inner, text="TOTAL ASIGNACIONES:", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['accent']).grid(row=len(asignaciones)+1, column=0, sticky='w', pady=(4,4))
        top.lbl_total_asig = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 9, 'bold'),
                                      bg=self.colors['card_bg'], fg=self.colors['success'])
        top.lbl_total_asig.grid(row=len(asignaciones)+1, column=1, sticky='e', pady=(4,4))

        ttk.Separator(results_inner, orient='horizontal').grid(row=len(asignaciones)+2, column=0, columnspan=2, sticky='ew', pady=4)

        row_offset = len(asignaciones) + 3
        tk.Label(results_inner, text="DEDUCCIONES", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['error']).grid(row=row_offset, column=0, columnspan=2, sticky='w', pady=(0,4))

        deducciones = [
            ("IVSS (trabajador)", "0,00"),
            ("FAOV", "0,00"),
            ("INCES", "0,00"),
        ]
        labels_ded = {}
        for i, (text, _) in enumerate(deducciones):
            tk.Label(results_inner, text=f"{text}:", font=('Segoe UI', 8),
                     bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=row_offset+i+1, column=0, sticky='w', pady=1)
            lbl = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 8, 'bold'),
                           bg=self.colors['card_bg'], fg=self.colors['error'])
            lbl.grid(row=row_offset+i+1, column=1, sticky='e', pady=1)
            labels_ded[text] = lbl

        tk.Label(results_inner, text="TOTAL DEDUCCIONES:", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['error']).grid(row=row_offset+len(deducciones)+1, column=0, sticky='w', pady=(4,4))
        top.lbl_total_ded = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 9, 'bold'),
                                     bg=self.colors['card_bg'], fg=self.colors['error'])
        top.lbl_total_ded.grid(row=row_offset+len(deducciones)+1, column=1, sticky='e', pady=(4,4))

        ttk.Separator(results_inner, orient='horizontal').grid(row=row_offset+len(deducciones)+2, column=0, columnspan=2, sticky='ew', pady=4)

        tk.Label(results_inner, text="NETO A PAGAR:", font=('Segoe UI', 11, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['accent']).grid(row=row_offset+len(deducciones)+3, column=0, sticky='w', pady=4)
        top.lbl_neto = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 12, 'bold'),
                                bg=self.colors['card_bg'], fg=self.colors['accent'])
        top.lbl_neto.grid(row=row_offset+len(deducciones)+3, column=1, sticky='e', pady=4)

        ttk.Separator(results_inner, orient='horizontal').grid(row=row_offset+len(deducciones)+4, column=0, columnspan=2, sticky='ew', pady=4)

        row_offset2 = row_offset + len(deducciones) + 5
        tk.Label(results_inner, text="APORTES PATRONALES", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['warning']).grid(row=row_offset2, column=0, columnspan=2, sticky='w', pady=(0,4))

        patronales = [
            ("IVSS patronal", "0,00"),
            ("LOPCYMAT", "0,00"),
        ]
        labels_pat = {}
        for i, (text, _) in enumerate(patronales):
            tk.Label(results_inner, text=f"{text}:", font=('Segoe UI', 8),
                     bg=self.colors['card_bg'], fg=self.colors['text_gray']).grid(row=row_offset2+i+1, column=0, sticky='w', pady=1)
            lbl = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 8, 'bold'),
                           bg=self.colors['card_bg'], fg=self.colors['warning'])
            lbl.grid(row=row_offset2+i+1, column=1, sticky='e', pady=1)
            labels_pat[text] = lbl

        tk.Label(results_inner, text="TOTAL APORTES PATRONALES:", font=('Segoe UI', 9, 'bold'),
                 bg=self.colors['card_bg'], fg=self.colors['warning']).grid(row=row_offset2+len(patronales)+1, column=0, sticky='w', pady=(4,4))
        top.lbl_total_pat = tk.Label(results_inner, text="Bs 0,00", font=('Segoe UI', 9, 'bold'),
                                     bg=self.colors['card_bg'], fg=self.colors['warning'])
        top.lbl_total_pat.grid(row=row_offset2+len(patronales)+1, column=1, sticky='e', pady=(4,4))

        top.labels = {**labels_asig, **labels_ded, **labels_pat}
        top.resultados = {}

        btn_frame = tk.Frame(main_frame, bg=self.colors['bg_main'])
        btn_frame.pack(fill='x', pady=8)

        self._crear_boton(btn_frame, "EXPORTAR A EXCEL", 
                         lambda: self._exportar_calculo_excel(top), 'info').pack(side='left', padx=4)
        self._crear_boton(btn_frame, "CERRAR", top.destroy, 'danger').pack(side='left', padx=4)

    def _procesar_calculo_nomina(self, top, combo_trab, periodo_var, salario_minimo_var,
                                 cestaticket_diario_var, lopcymat_porc_var,
                                 ivss_patronal_porc_var, ivss_trab_porc_var,
                                 faov_porc_var, inces_porc_var):
        
        seleccion = combo_trab.get()
        if not seleccion:
            self._mostrar_mensaje("Seleccione un trabajador", 'warning')
            return
        cedula = seleccion.split(" - ")[0]
        trabajador = self.nomina.trabajadores.get(cedula)
        if not trabajador:
            self._mostrar_mensaje("Trabajador no encontrado", 'error')
            return

        periodo = periodo_var.get()
        dias = 15 if periodo == "Quincenal" else 30

        hoy = datetime.now().date()
        fecha_inicio = hoy - timedelta(days=dias)
        fecha_inicio_str = fecha_inicio.strftime("%Y-%m-%d")
        fecha_hoy_str = hoy.strftime("%Y-%m-%d")

        registros = self.nomina.obtener_produccion(cedula)
        registros_periodo = [r for r in registros if fecha_inicio_str <= r.fecha <= fecha_hoy_str]
        total_produccion = sum(r.total for r in registros_periodo)

        salario_produccion = total_produccion
        salario_diario = salario_produccion / dias if dias > 0 else 0

        alicuota_utilidades_diaria = (salario_diario * 15) / 360
        alicuota_bono_diaria = (salario_diario * 15) / 360
        alicuota_utilidades = alicuota_utilidades_diaria * dias
        alicuota_bono = alicuota_bono_diaria * dias

        salario_integral_diario = salario_diario + alicuota_utilidades_diaria + alicuota_bono_diaria
        salario_integral_periodo = salario_integral_diario * dias

        try:
            valor_cestaticket_diario = float(cestaticket_diario_var.get().replace(',', '.'))
        except:
            valor_cestaticket_diario = 15.0
        cestaticket = valor_cestaticket_diario * dias

        total_asignaciones = salario_produccion + cestaticket + alicuota_utilidades + alicuota_bono

        try:
            salario_minimo = float(salario_minimo_var.get().replace(',', '.'))
        except:
            salario_minimo = 130.0

        tope_mensual = 5 * salario_minimo
        tope_periodo = tope_mensual * (dias / 30)

        try:
            porc_ivss_trab = float(ivss_trab_porc_var.get().replace(',', '.')) / 100
        except:
            porc_ivss_trab = 0.04
        base_ivss = min(salario_produccion, tope_periodo)
        ivss_trab = base_ivss * porc_ivss_trab

        try:
            porc_faov = float(faov_porc_var.get().replace(',', '.')) / 100
        except:
            porc_faov = 0.01
        faov = salario_integral_periodo * porc_faov

        try:
            porc_inces = float(inces_porc_var.get().replace(',', '.')) / 100
        except:
            porc_inces = 0.005
        inces = salario_produccion * porc_inces

        total_deducciones = ivss_trab + faov + inces
        neto = total_asignaciones - total_deducciones

        try:
            porc_ivss_pat = float(ivss_patronal_porc_var.get().replace(',', '.')) / 100
        except:
            porc_ivss_pat = 0.11
        ivss_patronal = salario_produccion * porc_ivss_pat

        try:
            porc_lopcymat = float(lopcymat_porc_var.get().replace(',', '.')) / 100
        except:
            porc_lopcymat = 0.01
        lopcymat = salario_produccion * porc_lopcymat

        total_patronales = ivss_patronal + lopcymat

        top.resultados = {
            'trabajador': trabajador.nombre,
            'cedula': cedula,
            'cargo': trabajador.cargo,
            'periodo': periodo,
            'dias': dias,
            'fecha_inicio': fecha_inicio_str,
            'fecha_fin': fecha_hoy_str,
            'salario_produccion': salario_produccion,
            'cestaticket': cestaticket,
            'alicuota_utilidades': alicuota_utilidades,
            'alicuota_bono': alicuota_bono,
            'total_asignaciones': total_asignaciones,
            'ivss_trab': ivss_trab,
            'faov': faov,
            'inces': inces,
            'total_deducciones': total_deducciones,
            'neto': neto,
            'ivss_patronal': ivss_patronal,
            'lopcymat': lopcymat,
            'total_patronales': total_patronales,
            'salario_diario': salario_diario,
            'salario_integral_diario': salario_integral_diario,
            'base_ivss': base_ivss,
            'tope_periodo': tope_periodo,
            'valor_cestaticket_diario': valor_cestaticket_diario,
        }

        labels = top.labels
        labels["Salario de producción"].config(text=self._formato_cop(salario_produccion))
        labels["Cestaticket"].config(text=self._formato_cop(cestaticket))
        labels["Alícuota utilidades"].config(text=self._formato_cop(alicuota_utilidades))
        labels["Alícuota bono vacacional"].config(text=self._formato_cop(alicuota_bono))
        top.lbl_total_asig.config(text=self._formato_cop(total_asignaciones))

        labels["IVSS (trabajador)"].config(text=self._formato_cop(ivss_trab))
        labels["FAOV"].config(text=self._formato_cop(faov))
        labels["INCES"].config(text=self._formato_cop(inces))
        top.lbl_total_ded.config(text=self._formato_cop(total_deducciones))

        top.lbl_neto.config(text=self._formato_cop(neto))

        labels["IVSS patronal"].config(text=self._formato_cop(ivss_patronal))
        labels["LOPCYMAT"].config(text=self._formato_cop(lopcymat))
        top.lbl_total_pat.config(text=self._formato_cop(total_patronales))

        self._mostrar_mensaje("Cálculo completado", 'success')

    def _exportar_calculo_excel(self, top):
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment
        except ImportError:
            self._mostrar_mensaje("La librería openpyxl no está instalada. Instálala con: pip install openpyxl", 'error')
            return

        resultados = getattr(top, 'resultados', None)
        if not resultados:
            self._mostrar_mensaje("Primero debe procesar el cálculo.", 'warning')
            return

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar cálculo de nómina"
        )
        if not archivo:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Cálculo Nómina"

            bold_font = Font(bold=True)
            center_align = Alignment(horizontal='center')
            money_fmt = '#,##0.00 "Bs"'

            ws['A1'] = "INVERSPORT - CÁLCULO DE NÓMINA"
            ws['A1'].font = Font(size=14, bold=True)
            ws.merge_cells('A1:B1')
            ws['A2'] = f"Trabajador: {resultados['trabajador']} ({resultados['cedula']})"
            ws['A3'] = f"Cargo: {resultados['cargo']}"
            ws['A4'] = f"Período: {resultados['periodo']} - {resultados['dias']} días"
            ws['A5'] = f"Fechas: {resultados['fecha_inicio']} al {resultados['fecha_fin']}"

            row = 7
            ws.cell(row=row, column=1, value="ASIGNACIONES").font = bold_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            asignaciones = [
                ("Salario de producción", resultados['salario_produccion']),
                ("Cestaticket", resultados['cestaticket']),
                ("Alícuota utilidades", resultados['alicuota_utilidades']),
                ("Alícuota bono vacacional", resultados['alicuota_bono']),
            ]
            for desc, val in asignaciones:
                ws.cell(row=row, column=1, value=desc)
                ws.cell(row=row, column=2, value=val).number_format = money_fmt
                row += 1
            ws.cell(row=row, column=1, value="TOTAL ASIGNACIONES").font = bold_font
            ws.cell(row=row, column=2, value=resultados['total_asignaciones']).number_format = money_fmt
            row += 2

            ws.cell(row=row, column=1, value="DEDUCCIONES").font = bold_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            deducciones = [
                ("IVSS (trabajador)", resultados['ivss_trab']),
                ("FAOV", resultados['faov']),
                ("INCES", resultados['inces']),
            ]
            for desc, val in deducciones:
                ws.cell(row=row, column=1, value=desc)
                ws.cell(row=row, column=2, value=val).number_format = money_fmt
                row += 1
            ws.cell(row=row, column=1, value="TOTAL DEDUCCIONES").font = bold_font
            ws.cell(row=row, column=2, value=resultados['total_deducciones']).number_format = money_fmt
            row += 2

            ws.cell(row=row, column=1, value="NETO A PAGAR").font = bold_font
            ws.cell(row=row, column=2, value=resultados['neto']).number_format = money_fmt
            row += 2

            ws.cell(row=row, column=1, value="APORTES PATRONALES").font = bold_font
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            row += 1
            patronales = [
                ("IVSS patronal", resultados['ivss_patronal']),
                ("LOPCYMAT", resultados['lopcymat']),
            ]
            for desc, val in patronales:
                ws.cell(row=row, column=1, value=desc)
                ws.cell(row=row, column=2, value=val).number_format = money_fmt
                row += 1
            ws.cell(row=row, column=1, value="TOTAL APORTES PATRONALES").font = bold_font
            ws.cell(row=row, column=2, value=resultados['total_patronales']).number_format = money_fmt

            row += 2
            ws.cell(row=row, column=1, value="Bases de cálculo").font = bold_font
            row += 1
            bases = [
                ("Salario diario", resultados['salario_diario']),
                ("Salario integral diario", resultados['salario_integral_diario']),
                ("Tope periódico (5 SM)", resultados['tope_periodo']),
                ("Base IVSS", resultados['base_ivss']),
                ("Cestaticket diario", resultados['valor_cestaticket_diario']),
            ]
            for desc, val in bases:
                ws.cell(row=row, column=1, value=desc)
                ws.cell(row=row, column=2, value=val).number_format = money_fmt
                row += 1

            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 18

            wb.save(archivo)
            self._mostrar_mensaje(f"Cálculo exportado a {archivo}", 'success')
        except Exception as e:
            self._mostrar_mensaje(f"Error al exportar: {str(e)}", 'error')